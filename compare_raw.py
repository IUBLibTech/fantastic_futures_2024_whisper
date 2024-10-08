#!/usr/bin/env python3

import argparse
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from transcript_compare import compare_transcripts
from transcript_spreadsheet import apply_function, expand_arguments, normalize_sheet_title, render_sheet
from utils import human_time
from transcript_loader import load_transcripts
import itertools
import pprint
import copy

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", default=False, action="store_true", help="Turn on debugging")
    parser.add_argument("threeplay", type=Path, help="3play json directory")
    parser.add_argument("workdir", type=Path, help="Root of the media files")
    parser.add_argument("output", type=Path, help="Excel Output File")
    
    args = parser.parse_args()
    logging.basicConfig(level=logging.DEBUG if args.debug else logging.INFO,
                        format="%(asctime)s [%(process)d:%(filename)s:%(lineno)d] [%(levelname)s] %(message)s")
    workdir: Path = args.workdir
    threeplay: Path = args.threeplay

    # Load all of the generated transcripts into a standard format.
    transcripts = []
    for d in workdir.glob('*'):
        transcripts.extend(load_transcripts(d, threeplay))

    # Do the comparison computation.
    variations = {'Whisper Model': ['small', 'medium', 'large-v2', 'large-v3'],
                  'Audio Filter': ['X', 'A', 'B'],
                  'Previous Text': ['T', 'F']}



    data = compute_raw_transcript_data(variations, transcripts, differences=True)
    #pprint.pprint(data)
    #exit(1)

    # Start generating the excel workbook
    workbook = Workbook()
    sheet: Worksheet = workbook.get_sheet_by_name("Sheet")
    columns = ['title', 'filename', 'content_type', 'physical_format', 'model', 
               'filter', 'previous_text', 'original_duration', 'truncated_duration',
               'whisper_processing_duration', 'processing_ratio', 'wer', 'wil', 'wip', 'mer', 
               'hits', 'substitutions', 'insertions', 'deletions']    
    row = 1
    for c, header in enumerate(columns):
        sheet.cell(row, c + 1, header)
    for d in data:
        row += 1
        for c, header in enumerate(columns):
            sheet.cell(row, c + 1, d.get(header, ''))
    workbook.save(args.output)

 
def group_by(transcripts: list[dict], field: str) -> dict[str, list[dict]]:
    """Return a dict of lists grouped by the given field"""
    results = {}
    for i in transcripts:
        k = i.get(field, None)
        if k not in results:
            results[k] = []
        results[k].append(i)
    return results


def search_transcripts(transcripts: list[dict], query: dict) -> list[dict]:
    """Search the transcripts for the specific query fields"""
    results = list(transcripts)
    for k, v in query.items():
        results = [x for x in results if x.get(k, None) == v]
    return results



def compute_raw_transcript_data(variations: dict[str, list], transcripts: list[dict], differences=True,
                            edit_width=75):
    """Gather all of the transcript comparison data for each worksheet"""
    worksheets = []
    for title, txscripts in sorted(group_by(transcripts, 'title').items()):
        for file, fxscripts in group_by(txscripts, 'base_filename').items():            
            sheet = {'title': title,
                     'filename': file,
                     'content_type': fxscripts[0]['content_type'],
                     'physical_format': fxscripts[0]['physical_format'],
                     'original_duration': fxscripts[0]['original_duration'],
                     'truncated_duration': fxscripts[0]['truncated_duration'],
                     }
            for perm in list(itertools.product(*variations.values())):
                res = search_transcripts(fxscripts, query={'variant': perm})
                if len(res) != 1:
                    logging.warning(f"Skipping {perm}: not found for {file}")                    
                    continue
                data = {'processing_ratio':  res[0]['processing_ratio'],
                        'whisper_processing_duration': res[0]['whisper_processing_duration'],
                        'model': perm[0],
                        'filter': perm[1],
                        'previous_text': perm[2]}
                data.update(compare_transcripts(res[0]['3play_transcript'],
                                                res[0]['whisper_transcript'],
                                                gen_viz=False))
                
                xdata = {}
                xdata.update(sheet)
                xdata.update(data)
                key = (title, *perm)
                worksheets.append(xdata)


    return worksheets




if __name__ == "__main__":
    main()