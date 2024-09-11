#!/usr/bin/env python3

import argparse
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from transcript_compare import compare_transcripts
from transcript_spreadsheet import apply_function, expand_arguments, normalize_sheet_title, render_sheet
from utils import human_time
from transcript_loader import load_transcripts
import itertools
import pprint
import copy

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", default=False, action="store_true", help="Turn on debugging")
    parser.add_argument("threeplay", type=Path, help="3play json directory")
    parser.add_argument("workdir", type=Path, help="Root of the media files")
    parser.add_argument("output", type=Path, help="Excel Output File")
    
    args = parser.parse_args()
    logging.basicConfig(level=logging.DEBUG if args.debug else logging.INFO,
                        format="%(asctime)s [%(process)d:%(filename)s:%(lineno)d] [%(levelname)s] %(message)s")
    workdir: Path = args.workdir
    threeplay: Path = args.threeplay

    # Load all of the generated transcripts into a standard format.
    transcripts = []
    for d in workdir.glob('*'):
        transcripts.extend(load_transcripts(d, threeplay))

    # Do the comparison computation.
    variations = {'Whisper Model': ['small', 'medium', 'large-v2', 'large-v3'],
                  'Audio Filter': ['X', 'A', 'B'],
                  'Previous Text': ['T', 'F']}
    data = compute_transcript_data(variations, transcripts, differences=True)

    # Start generating the excel workbook
    workbook = Workbook()
    del(workbook['Sheet']) # remove the initial sheet.

    row_defs = [None,
                ['Processing Ratio', 'processing_ratio', '0.00%'],
                None,
                ['Word Error Rate', 'wer', '0.00%'],
                ['Word Information Lost', 'wil', '0.00%'],
                ['Word Information Preserved', 'wip', '0.00%'],
                ['Match Error Rate', 'mer', '0.00%'],
                None,
                ['Hits', 'hits', '0'],
                ['Substitutions', 'substitutions', '0'],
                ['Insertions', 'insertions', '0'],
                ['Deletions', 'deletions', '0'],
                None,
                ['Edits', 'visualization', None]]
    for k, v in data.items():
        render_sheet(workbook, k, v,
                     variations, row_defs)

    # for the overall average sheet, the field should be the average function for everything    
    for r in row_defs:
        if r is not None:
            r[1] = apply_function('AVERAGE', all_sheets := set(data.keys()))
    # remove all but the first 8 rows, since we don't care about edits or the
    # individual hit/sub/del/ins.
    row_defs = row_defs[0:8]
    
    avg_sheet = render_sheet(workbook, "Average", {'title': "Average Across Everything",
                                                   'filename': '',
                                                   'truncated_duration': (total_truncated_duration := sum([x['truncated_duration'] for x in data.values()])),
                                                   'original_duration': (total_original_duration := sum([x['original_duration'] for x in data.values()])),
                                                   'physical_format': 'All Physical Formats',
                                                   'content_type': "All Content Types"},
                                       variations, row_defs, edit_width=10, position=0)

    # we need to put some other stats in the Average page.  These should be
    # genericized, but I'm pressed for time.
    row = 15    
    col_max = len(list(itertools.product(*variations.values())))
    # Previous Text Deltas
    avg_sheet.cell(row, 1, "Previous Text Deltas")
    this_cell = expand_arguments(("Average",), (-5, 0))[0]
    left_cell = expand_arguments(("Average",), (-5, -1))[0]
    for col in range(0, col_max, 2):
        avg_sheet.cell(row, col + 3, f"={this_cell}-{left_cell}")
        
    # Audio Filter Deltas
    row += 1
    row_offset_base = -6
    for delta, column_offset, start_offset in (('Audio Filter Delta A', -2, 4),  
                                               ('Audio Filter Delta B', -4, 6)):
        for i, val in enumerate(['T', 'F']):
            avg_sheet.cell(row, 1, f"{delta}/{val}")
            for col in range(0, col_max, 6):
                this_cell = expand_arguments(("Average",), (row_offset_base, 0))[0]
                left_cell = expand_arguments(("Average",) , (row_offset_base, column_offset))[0]
                avg_sheet.cell(row, col + start_offset + i, f"={this_cell} - {left_cell}")
            row+= 1
            row_offset_base -= 1

 

    # for each of the subset aggregates, we want the average WER data for the set
    # but we also want them for the others.  So add a header for "not including this"
    # and repeat the WER fields again...
    row_defs.append(['Not including this subset', None, None])
    row_defs.extend(copy.deepcopy(row_defs[3:8]))

    for category, key in (('Physical Format', 'physical_format'),
                          ('Content Category', 'content_type')):
        for cvalue in reversed(sorted(set([x[key] for x in data.values()]))): 
            csheets = set([x for x in data.keys() if data[x][key] == cvalue])
            # apply the average function in place for row_defs 0-7
            for i in range(0, 8):
                if row_defs[i] is not None:
                    row_defs[i][1] = apply_function("AVERAGE", csheets)
            # apply the average function adjusted by -6 for rows 9-12
            for i in range(9, 13):
                if row_defs[i] is not None:
                    row_defs[i][1] = apply_function("AVERAGE", all_sheets - csheets, (-6, 0))

            subset_truncated_duration = 0
            subset_original_duration = 0
            for s in csheets:
                subset_truncated_duration += data[s]['truncated_duration']
                subset_original_duration += data[s]['original_duration']

            # render the sheet.
            render_sheet(workbook, normalize_sheet_title(f"{category} - {cvalue}"), 
                         {'title': f"{category} - {cvalue}",
                          'filename': '',
                          'truncated_duration': subset_truncated_duration,
                          'original_duration': subset_original_duration,
                          'physical_format': '',
                          'content_type': ''},
                          variations, row_defs, edit_width=10, position=1)

    workbook.save(args.output)

 
def group_by(transcripts: list[dict], field: str) -> dict[str, list[dict]]:
    """Return a dict of lists grouped by the given field"""
    results = {}
    for i in transcripts:
        k = i.get(field, None)
        if k not in results:
            results[k] = []
        results[k].append(i)
    return results


def search_transcripts(transcripts: list[dict], query: dict) -> list[dict]:
    """Search the transcripts for the specific query fields"""
    results = list(transcripts)
    for k, v in query.items():
        results = [x for x in results if x.get(k, None) == v]
    return results


def compute_transcript_data(variations: dict[str, list], transcripts: list[dict], differences=True,
                            edit_width=75):
    """Gather all of the transcript comparison data for each worksheet"""
    worksheets = {}
    for title, txscripts in sorted(group_by(transcripts, 'title').items()):
        sheet_num = 0
        for file, fxscripts in group_by(txscripts, 'base_filename').items():            
            sheet = {'title': title,
                     'filename': file,
                     'content_type': fxscripts[0]['content_type'],
                     'physical_format': fxscripts[0]['physical_format'],
                     'original_duration': fxscripts[0]['original_duration'],
                     'truncated_duration': fxscripts[0]['truncated_duration'],
                     'variations': {} 
                     }
            for perm in list(itertools.product(*variations.values())):
                res = search_transcripts(fxscripts, query={'variant': perm})
                if len(res) != 1:
                    logging.warning(f"Skipping {perm}: not found for {file}")                    
                    continue
                data = {'processing_ratio':  res[0]['processing_ratio'],
                        'whisper_processing_duration': res[0]['whisper_processing_duration']}
                data.update(compare_transcripts(res[0]['3play_transcript'],
                                                res[0]['whisper_transcript'],
                                                edit_width, differences))
                sheet['variations'][perm] = data

            sheet_num += 1
            worksheet_name = normalize_sheet_title(title) + f"_{sheet_num}"
            worksheets[worksheet_name] = sheet
    return worksheets


if __name__ == "__main__":
    main()