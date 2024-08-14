#!/bin/env python3
"""
Populate a source_content directory with MDPI content based on 
* A file which has the HCP S3 object names
* An inventory spreadsheet where the MDPI column is a comma separated list of
    barcodes in display order
* each item has a metadata.yaml file with the spreadsheet row it corresponds to.
"""
import argparse
import boto3
import logging
from openpyxl.reader.excel import load_workbook
from pathlib import Path
import re
import yaml

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--debug", default=False, action="store_true", help="Turn on debugging")
    parser.add_argument("s3_config", help="S3 Configuration file")
    parser.add_argument("s3_list", help="List of objects in the s3 bucket")
    parser.add_argument("project_inventory", help="Project Inventory Spreadsheet")
    parser.add_argument("destination", type=Path, help="Destination root for the media files")
    args = parser.parse_args()
    logging.basicConfig(format="%(asctime)s [%(levelname)-8s] (%(filename)s:%(lineno)d:%(process)d)  %(message)s",
                        level=logging.DEBUG if args.debug else logging.INFO)

    if not args.destination.is_dir():
        logging.error("The destination must be a directory")
        exit(1)

    # load the S3 configuration
    with open(args.s3_config) as f:
        s3_config = yaml.safe_load(f)
    s3 = boto3.client('s3',
                      aws_access_key_id=s3_config['id'],
                      aws_secret_access_key=s3_config['key'],
                      endpoint_url=f"https://{s3_config['host']}")

    # load our S3 file list
    s3_list = load_s3_list(args.s3_list)

    # load the project inventory
    inventory = read_inventory(args.project_inventory)

    destination: Path = args.destination
    for title, data in inventory.items():        
        ctitle = normalize_title(title)
        tpath = destination / ctitle
        logging.info(tpath)        
        tpath.mkdir(exist_ok=True)
        # drop the row metadata here
        with open(tpath / "metadata.yaml", "w") as f:
            yaml.safe_dump(data, f)
        if data['mdpi'] is not None:
            bcount = 1
            for barcode in data['mdpi']:
                scount = 1
                for sfile in s3_list[barcode]:
                    nfile = tpath / f"{bcount:02d}-{scount:02d}-{barcode}.mp4"
                    logging.info(f"{sfile} -> {nfile}")                    
                    s3.download_file(s3_config['bucket'], sfile, str(nfile))                    
                    scount += 1
                bcount += 1


def read_inventory(file: Path) -> dict:
    """Read the inventory spreadsheet"""
    logging.debug("Reading inventory")
    workbook = load_workbook(file)
    # the sheet with our info is on "File list only"
    #sheet = workbook.worksheets[0]
    sheet = workbook["File list only"]
    # read the first row to get the column titles.
    column_indexes = {}
    for c in range(1, sheet.max_column + 1):
        x = str(sheet.cell(1, c).value).lower()
        if x:
            column_indexes[x] = c

    # make a map between the title and the row
    res = {}
    for r in range(2, sheet.max_row + 1):
        row = {}
        for name, c in column_indexes.items():
            v = sheet.cell(r, c).value
            row[name] = None if v is None else str(v)
        if 'mdpi' in row and row['mdpi'] is not None:
            row['mdpi'] = row['mdpi'].split(',')
        if row.get('title', None) is not None:
            res[row['title']] = row
    return res


def load_s3_list(file: Path) -> dict:
    """Load the S3 inventory file, keyed by MDPI barcode and only containing 
    high quality derivatives"""
    logging.debug("Loading S3 List")
    res = {}
    with open(file) as f:
        for line in f.readlines():
            line = line.strip()
            if "_MDPI_" not in line:
                # not an MDPI thing.
                continue
            if "_high_" not in line:
                # not a high quality derivative
                continue
            barcode=line.split('_')[2]
            if barcode not in res:
                res[barcode] = []
            res[barcode].append(line)
    return res


def normalize_title(text):
    """Normalize title text so it can be a filesystem name"""
    res = re.sub(r'[^A-Za-z0-9\.\-]', '_', text.strip(), re.X)
    res = re.sub(r'_+', '_', res)
    return res.strip('_.-')


if __name__ == "__main__":
    main()