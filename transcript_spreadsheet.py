import itertools
import string

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.utils.cell import get_column_letter

from utils import human_time
import logging

def render_sheet(workbook: Workbook, sheet_title: str, data: dict,
                 variations: dict, row_defs: list, edit_width=75,
                 position=-1):
    sheet: Worksheet = workbook.create_sheet(sheet_title, position)

    # Generate the main title bit at the first row on the 2nd column
    text_title = f"{data['title']} {data['filename']} ({human_time(data['truncated_duration'])} / {human_time(data['original_duration'])})"
    sheet.cell(1, 2, text_title).font = Font(bold=True, name="Arial", sz=14)
    # and the physical media & content type on the 2nd row.
    for i, k in enumerate(['physical_format', 'content_type']):
        sheet.cell(2, i + 2, data[k])

    data_font = Font(name="Arial", sz=10)
    edit_font = Font(name="Courier", sz=8)
    vtitle_font = Font(name="Arial", sz=10, bold=True)

    # walk through the variation titles & and the row definitions
    top_row = row = 4
    maxlen = 0
    for rowdef in (*list(variations.keys()), *row_defs):
        if rowdef is not None:
            if isinstance(rowdef, str):
                sheet.cell(row, 1, rowdef).font = data_font
                maxlen = max(maxlen, len(rowdef))
            elif isinstance(rowdef, (list, tuple)):
                sheet.cell(row, 1, rowdef[0]).font = data_font
                maxlen = max(maxlen, len(rowdef[0]))
        row += 1
    sheet.column_dimensions['A'].width = maxlen

    # walk through the permutations!
    col = 1
    for perm in list(itertools.product(*variations.values())):
        col += 1
        row = top_row
        for p in perm:
            sheet.cell(row, col, p).font = vtitle_font
            row += 1
        
        sheet.column_dimensions[get_column_letter(col)].width = edit_width  # chr(64 + col)
        for rowdef in row_defs:
            if rowdef is not None:
                if len(rowdef) == 3:
                    _, field, format = rowdef
                    if field is None:
                        pass
                    elif field.startswith('='):
                        # this is a function.
                        c = sheet.cell(row, col, field)
                        c.font = data_font
                        c.number_format = format
                    else:
                        # this is a constant
                        if perm in data['variations']:
                            fdata = data['variations'][perm][field]
                            if isinstance(fdata, list):
                                # this only applies to the edits...
                                for l in fdata:
                                    c = sheet.cell(row, col, l)                            
                                    c.font = edit_font
                                    row += 1                          
                            else:                  
                                c = sheet.cell(row, col, fdata)
                                c.font = data_font
                                c.number_format = format


            row += 1
    return sheet

def apply_function(function: str, sheets: set, offset=(0, 0)):
    """Create an excel function with the arguments which span
       across all of the specified sheets for the current cell
       with an optional offset"""
    return f"={function}({','.join(expand_arguments(sheets, offset))})"


def expand_arguments(sheets: set, offset=(0, 0)):
    """Return a list of arguments that reference the current cell with an optional
       offset across all of the specified sheets"""
    result = []
    for s in sheets:
        result.append(f'INDIRECT(ADDRESS(ROW()+{offset[0]}, COLUMN()+{offset[1]},,,"{s}"))')
    return result


def normalize_sheet_title(text: str) -> str:
    """Remove invalid characters from sheet titles"""
    ntext = ""
    cksum = 0
    for c in text:
        cksum += ord(c)
        if c in string.ascii_letters or c in string.digits:
            ntext += c
    if len(ntext) > 25:
        ntext = ntext[0:25]
    ntext += chr(65 + (cksum % 24))
    return ntext
